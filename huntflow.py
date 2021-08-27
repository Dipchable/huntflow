import requests
import openpyxl
import os
import glob
import unicodedata
from requests_toolbelt.multipart.encoder import MultipartEncoder

wb = openpyxl.load_workbook('Тестовая база.xlsx')
sheet = wb.active

now_dir = os.getcwd()

for i in range(2, 3):
	if os.path.isdir(sheet.cell(i,1).value):
		os.chdir(sheet.cell(i,1).value)
	file_name = unicodedata.normalize('NFKD', (sheet.cell(i,2).value))
	file_name = glob.glob(f"{file_name.strip()}.*")
	m = MultipartEncoder(fields={'file':('Глибин Виталий Николаевич.doc', open(file_name[0], 'rb'))})
	files = {'file':(file_name[0], open(file_name[0], 'rb'))}
	url = 'https://dev-100-api.huntflow.dev/account/2/upload'
	response = requests.post(url, headers={'Content-Type':m.content_type, 'User-Agent': 'App/1.0','X-File-Parse': 'true','Authorization': 'Bearer <token>'}, files=files)
	print(response.content)
	os.chdir('..')